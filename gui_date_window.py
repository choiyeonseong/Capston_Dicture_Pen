

import sys
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QComboBox
from openpyxl import load_workbook

from gui_table_window import *


class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.lbl = QLabel('날짜 선택', self)
        self.lbl.setAlignment(Qt.AlignCenter)

        load_wb = load_workbook("resource/dictionary_example.xlsx", data_only=True)
        lists = load_wb.sheetnames
        cb = QComboBox(self)
        cb.addItems(lists)

        cb.activated[str].connect(self.onActivated)

        layout=QVBoxLayout()
        layout.addWidget(self.lbl)
        layout.addWidget(cb)

        self.setLayout(layout)

        self.setWindowTitle('날짜 선택')
        # self.setGeometry(1000, 500, 100, 100)
        self.show()

        # print(cb.currentText())

    def onActivated(self, text):

        self.tableWindow = tablewindow(text)
        self.tableWindow.setWindowTitle(text)
        self.tableWindow.resize(1000, 400)
        self.tableWindow.show()



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())